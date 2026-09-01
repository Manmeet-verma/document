"""Excel workbook generation with formatting and image hyperlinks."""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .extractor import AADHAAR_COLUMNS, RC_COLUMNS, NOT_DETECTED, VERIF_REQ, field_value

HEADER_COLOR = "1F4E78"
ALT_ROW_COLOR = "DDEBF7"
WARN_COLOR = "FFF2CC"
OK_COLOR = "E2EFDA"

COLUMNS: list[tuple[str, int]] = [
    ("Folder / Person ID", 22),
    ("Aadhaar Images", 40),
    ("RC Images", 40),
] + [(h, 26) for _, h in AADHAAR_COLUMNS] + [(h, 26) for _, h in RC_COLUMNS]


def _status_for(field) -> str:
    if field is None:
        return "not detected"
    return field.get("status", "not detected")


def generate_excel(rows: list[dict], output_path: str, image_base: str = "") -> str:
    """Write the final formatted Excel file. Returns the output path."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)

    for col_idx, (header, _width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = _width

    for r_idx, row in enumerate(rows, start=2):
        person_id = row.get("person_id", "")
        fill = PatternFill("solid", fgColor=ALT_ROW_COLOR) if r_idx % 2 == 0 else None

        c = ws.cell(row=r_idx, column=1, value=person_id)
        c.alignment = Alignment(vertical="top")

        self_dir = image_base or ""
        _write_image_cell(ws, r_idx, 2, row.get("aadhaar_images", ""), self_dir)
        _write_image_cell(ws, r_idx, 3, row.get("rc_images", ""), self_dir)

        col_idx = 4
        for key, _header in AADHAAR_COLUMNS:
            _write_field_cell(ws, r_idx, col_idx, row.get(key))
            col_idx += 1
        for key, _header in RC_COLUMNS:
            _write_field_cell(ws, r_idx, col_idx, row.get(key))
            col_idx += 1

        if fill:
            for j in range(1, len(COLUMNS) + 1):
                ws.cell(row=r_idx, column=j).fill = fill

    ws.freeze_panes = "A2"
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(COLUMNS)):
        for cell in row_cells:
            cell.border = _thin_border()

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

    wb.save(output_path)
    return output_path


def _write_image_cell(ws, r_idx: int, c_idx: int, image_list: str, base: str):
    center = Alignment(vertical="center", horizontal="left", wrap_text=True)
    if not image_list:
        ws.cell(row=r_idx, column=c_idx, value=NOT_DETECTED).alignment = center
        return
    paths = [p for p in image_list.split("\n") if p]
    cell = ws.cell(row=r_idx, column=c_idx)
    lines = []
    for p in paths:
        name = os.path.basename(p)
        full = os.path.join(base, p) if base else p
        target = full if os.path.exists(full) else p
        cell.hyperlink = f"file:///{target.replace(os.sep, '/').lstrip('/')}"
        lines.append(f"Open: {name}")
    cell.value = "\n".join(lines)
    cell.alignment = center
    cell.font = Font(color="0563C1", underline="single")


def _write_field_cell(ws, r_idx: int, c_idx: int, field):
    value = _status_for(field)
    cell = ws.cell(row=r_idx, column=c_idx)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    if value == "not detected":
        cell.value = NOT_DETECTED
        cell.fill = PatternFill("solid", fgColor=WARN_COLOR)
    elif value == "verification required":
        cell.value = VERIF_REQ
        cell.fill = PatternFill("solid", fgColor=WARN_COLOR)
    else:
        cell.value = field_value(field)
        cell.fill = PatternFill("solid", fgColor=OK_COLOR)
    return cell


def _thin_border():
    side = Side(style="thin", color="B0B0B0")
    return Border(left=side, right=side, top=side, bottom=side)
